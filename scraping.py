
# import pandas lib as pd
# from importlib.metadata import requires
# import pandas as pd
from pickle import GLOBAL
from tkinter import font
from turtle import left, width
import openpyxl
# import requests
# from bs4 import BeautifulSoup
from selenium import webdriver
# from selenium.webdriver.common.by import By
from tkinter import *
from tkinter.ttk import *

import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo


fileURL = ""

searchResult = []

def fileScraping():
    global fileURL
    
    if fileURL == "":
        return
    
    global searchResult
    searchResult = []
    
    browser = webdriver.Chrome()

    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(fileURL)  
    
    # Define variable to read sheet
    dataframe1 = dataframe.active
    rows = dataframe1.iter_rows(1, dataframe1.max_row)
    cnt = 0
    
    for row in rows:
        resultString = "\""
        for col in range(0, 2):
            strArr = row[col].value.split()
            for indexStr in strArr:
                resultString += indexStr + "+"
        if cnt != 0:
            resultUrl = "https://duckduckgo.com/?q=" + resultString +"&t=h_&ia=web"
            browser.get(resultUrl)

            results = browser.find_element("id","links")
            
            data = results.text.split("\n")
            row[2].value = data[0]
            row[3].value = data[1]
            
            temp = []
            for i in range(0, 4):
                temp.append(row[i].value)
            searchResult.append(temp)
        cnt += 1
        
    dataframe.save("output.xlsx")
    
    if len(searchResult) > 0:
        # label_page['text'] = 1
        display()


def customScraping():
    if input_name.get() == "" and input_company.get() == "":
        return
    
    browser = webdriver.Chrome()
    
    resultUrl = "https://duckduckgo.com/?q=" + input_name.get() + input_company.get() +"&t=h_&ia=web"
    browser.get(resultUrl)
    
    results = browser.find_element("id","links")
    
    data = results.text.split("\n")
    
    text_result.delete("1.0", 'end')
    text_result.insert(END, data[0] + "\n\n" + data[1] + "\n\n" + data[2])
    

# scraping function
def scraping():
    if combo_searchType.get() == 'file':
        fileScraping()
    if combo_searchType.get() == 'input':
        customScraping()


# select file dialog function
def select_file():
    filetypes = (
        ('text files', '*.txt'),
        ('All files', '*.*')
    )

    filename = fd.askopenfilename(
        title='Open a file',
        initialdir='/',
        filetypes=filetypes)

    global fileURL 
    fileURL = filename
    
    label_filePath['text'] = fileURL

    showinfo(
        title='Selected File',
        message=filename
    )



def searchTypeChange(event):
    input_name.delete(0, END)
    input_company.delete(0, END)
    text_result.delete("1.0", 'end')
    
    if combo_searchType.get() == "file":
        root.geometry('1200x370')    
        btn_open.place(x=50, y=80)       
        label_filePath.place(x=130, y=82)
        frame.place(x=20, y=120)
        
        label_searchProperty_name.place_forget()
        label_searchProperty_company.place_forget()
        # input_name.place(x=50, y=150)
        # input_name.config(width=15)
        # input_company.place(x=50, y=180)
        # input_company.config(width=15)
        # text_result.place(x=150, y=150)
        # text_result.config(width=40, height=11)
        input_name.place_forget()
        input_company.place_forget()
        text_result.place_forget()
        
        # btn_next.place(x=280, y=120)
        # label_page.place(x=255, y=122)
        # btn_before.place(x=220, y=120)
        
        btn_run.place(x=500, y=80)
        btn_run.config(width=15)
        
    if combo_searchType.get() == 'input':
        global fileURL
        fileURL = ""
        label_filePath['text'] = ""
        display_table.delete(*display_table.get_children())
        
        # label_page['text'] = '0'
        
        root.geometry('500x500')
        btn_open.place_forget()
        label_filePath.place_forget()
        frame.place_forget()
        
        # btn_next.place_forget()
        # label_page.place_forget()
        # btn_before.place_forget()
        
        label_searchProperty_name.place(x=30, y=82)
        label_searchProperty_company.place(x=20, y=120)
        
        input_name.place(x=100, y=82)
        # input_name.config(width=60)
        input_company.place(x=100, y=120)
        # input_company.config(width=60)
        text_result.place(x=100, y=150)
        # text_result.config(width=45, height=15)
        
        btn_run.place(x=50, y=430)    
        btn_run.config(width=68)  
        
    return


# Dialog
root = tk.Tk()
root.title('Tkinter Open File Dialog')
root.resizable(False, False)
root.geometry('1200x370')



label_searchType = ttk.Label(
    root,
    text='Searching Type :',
    width=15
)
label_searchType.pack(expand=True)
label_searchType.place(x=30, y=30)


# combo box
combo_searchType = ttk.Combobox(
    root,
    text='file',
    width=10,
    textvariable='file')
combo_searchType['values'] = ['file', 'input']
combo_searchType['state'] = 'readonly'
combo_searchType.pack(expand=True)
combo_searchType.place(x=125, y=30)
combo_searchType.set('file')

combo_searchType.bind('<<ComboboxSelected>>', searchTypeChange)


# open button
btn_open = ttk.Button(
    root,
    text='Open a File',
    command=select_file
)
btn_open.pack(expand=True)
btn_open.place(x=50, y=80)


# label filepath
label_filePath = ttk.Label(
    root,
    text='',
    width=35,
    font=14,
    background ="lightgrey"
)
label_filePath.pack(expand=True)
label_filePath.place(x=130, y=82)


# input name
input_name = ttk.Entry(root, width=60)
input_name.pack(expand=True)
input_name.place(x=100, y=82)
input_name.place_forget()


# input company
input_company = ttk.Entry(root, width=60)
input_company.pack(expand=True)
input_company.place(x=100, y=120)
input_company.place_forget()


# label name
label_searchProperty_name = ttk.Label(
    root,
    text='Name',
    width=8,
)
label_searchProperty_name.pack(expand=True)
label_searchProperty_name.pack_forget()


# label company
label_searchProperty_company = ttk.Label(
    root,
    text='Company',
    width=10,
)
label_searchProperty_company.pack(expand=True)
label_searchProperty_company.pack_forget()

# def display():
#     global searchResult
    
#     input_name.delete(0, END)
#     input_company.delete(0, END)
#     text_result.delete("1.0", 'end')
    
#     input_name.insert(END, searchResult[int(label_page['text']) - 1][0])
#     input_company.insert(END, searchResult[int(label_page['text']) - 1][1])
#     text_result.insert(END, searchResult[int(label_page['text']) - 1][2] + "\n\n" + searchResult[int(label_page['text']) - 1][3])

# def next():
#     global searchResult
#     if label_page['text'] == len(searchResult) or searchResult == [] or label_page['text'] == 0:
#         return
    
#     label_page['text'] = int(label_page['text']) + 1
    
#     display()

# def before():
#     global searchResult
#     if label_page['text'] == 1 or label_page['text'] == 0:
#         return
    
#     label_page['text'] = int(label_page['text']) - 1
    
#     display()


# btn_next = ttk.Button(
#     root,
#     text='>',
#     width=2,
#     command=next
# )
# btn_next.pack()
# btn_next.place(x=280, y=120)


# label_page = ttk.Label(
#     root,
#     text='0',
#     width=1,
#     font=14,
# )
# label_page.pack(expand=True)
# label_page.place(x=255, y=122)


# btn_before = ttk.Button(
#     root,
#     text='<',
#     width=2,
#     command=before
# )
# btn_before.pack()
# btn_before.place(x=220, y=120)



# run button
btn_run = ttk.Button(
    root,
    text='Run Scraping',
    command=scraping,
    width=15
)
btn_run.pack()
btn_run.place(x=500, y=80)


text_result = Text(root, width=45, height=15)
text_result.pack()
text_result.place(x=100, y=150)
text_result.place_forget()


def display():
    for i in range(0, len(searchResult)):
        display_table.insert(parent='', index='end', iid=i, text='', 
            values=((i+1), searchResult[i][0], searchResult[i][1], searchResult[i][2], searchResult[i][3]))
    #     return
    # return


frame = Frame(root)
frame.pack()
frame.place(x=20, y=120)

display_table = ttk.Treeview(frame)

display_table['columns'] = ('item_id', 'item_name', 'item_company', 'item_position', 'item_url')

display_table.column("#0", width=0, stretch=NO)
display_table.column("item_id",anchor=CENTER, width=20)
display_table.column("item_name",anchor=W,width=120)
display_table.column("item_company",anchor=W,width=120)
display_table.column("item_position",anchor=W,width=450)
display_table.column("item_url",anchor=W,width=450)

display_table.heading("#0",text="",anchor=CENTER)
display_table.heading("item_id",text="Id",anchor=CENTER)
display_table.heading("item_name",text="Name",anchor=CENTER)
display_table.heading("item_company",text="Company",anchor=CENTER)
display_table.heading("item_position",text="Position",anchor=CENTER)
display_table.heading("item_url",text="Url",anchor=CENTER)

# display_table.insert(parent='',index='end',iid=0,text='',
# values=('1','Ninja','101','Oklahoma', 'Moore'))
# display_table.insert(parent='',index='end',iid=1,text='',
# values=('2','Ranger','102','Wisconsin', 'Green Bay'))
# display_table.insert(parent='',index='end',iid=2,text='',
# values=('3','Deamon','103', 'California', 'Placentia'))
# display_table.insert(parent='',index='end',iid=3,text='',
# values=('4','Dragon','104','New York' , 'White Plains'))
# display_table.insert(parent='',index='end',iid=4,text='',
# values=('5','CrissCross','105','California', 'San Diego'))
# display_table.insert(parent='',index='end',iid=5,text='',
# values=('6','ZaqueriBlack','106','Wisconsin' , 'TONY'))

display_table.pack()




# run the application
root.mainloop()